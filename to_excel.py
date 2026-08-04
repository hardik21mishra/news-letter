import pandas as pd
from datetime import date
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def export_to_excel(articles, filename=None):
    if filename is None:
        filename = f"news_today_{date.today().isoformat()}.xlsx"

    df = pd.DataFrame(articles)

    # Reorder columns for the curator
    df = df[["title", "source", "description", "published_date", "link", "summary"]]

    df["curator_notes"] = ""

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="News")

        worksheet = writer.sheets["News"]

        # Column widths — keyed by name, not hardcoded letters
        widths = {
            "title": 40,
            "source": 20,
            "description": 50,
            "published_date": 20,
            "link": 15,
            "summary": 40,
            "curator_notes": 30,
        }
        for col_name, width in widths.items():
            col_letter = get_column_letter(df.columns.get_loc(col_name) + 1)
            worksheet.column_dimensions[col_letter].width = width

        # Wrap text on all cells
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Freeze header row + autofilter
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        # Bold header row
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        # Clickable "Open article" links, derived from actual link column position
        link_col = get_column_letter(df.columns.get_loc("link") + 1)
        for i, row in enumerate(df.itertuples(), start=2):
            cell = worksheet[f"{link_col}{i}"]
            cell.hyperlink = row.link
            cell.value = "Open article"
            cell.style = "Hyperlink"

    return filename